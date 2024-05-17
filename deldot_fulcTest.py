import time
import requests
import json
from fulcrum import Fulcrum
from fulcrum.exceptions import NotFoundException, InternalServerErrorException
import sys
import argparse
from datetime import date
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
from io import StringIO

# https://web.fulcrumapp.com/records/636dd075-3d3a-4fd2-bab4-897700ce04a5
# https://web.fulcrumapp.com/dash/7760b3b1-8b7b-4724-a5e9-3b235f193fe7

# Get time the script started so you know how long it ran
start_time = time.time()

# Specify the full path to the CSV file
# csv_path = "Fulcrum/de_ada_curb_ramp_inventory.csv"

key = "0cff9ebe895325dcb1201a55c2b8c1b03a9e4214aaf3c4da7390e9297a300b8ccc77bce196c4abd0"

fulcrum = Fulcrum(key=key)
as_csv = fulcrum.query('SELECT * FROM "DE ADA Curb Ramp Inventory";', 'csv')
s=str(as_csv,'utf-8')
data = StringIO(s) 

# Read the CSV file into a pandas DataFrame
df = pd.read_csv(data)

# Define a function to filter columns based on prefix and absence of '_pic'
def filter_columns(prefix):
    return [col for col in df.columns if col.startswith(prefix) and '_pic' not in col and '_link' not in col]

project = 'US 202 Concord Pike'
df = df.loc[df['project'] == project]

# Define the prefix list of columns
# prefix_columns = ["_status", "_title", "date", "ramp_id", "project", "loc", "dir", "ramp_type"]
prefix_columns = ["date", "ramp_id", "loc", "dir", "ramp_type"]
# sufix_columns = ["drainage", "curb_hgt", "constraint", "pinch_wdth", "gap", "vert_diff", "height", "crosswalk", "stop_offst", "defic", "stan_det", "comp_recc", "review", "comments"]
sufix_columns = ["drainage", "curb_hgt", "constraint", "pinch_wdth", "gap", "vert_diff", "height", "crosswalk", "stop_offst", "review", "defic", "comments"]

# Drop records where ramp_type is null
df = df.dropna(subset=["ramp_type"])

# Create sub DataFrames for each type and add prefix list of columns
type_1a_df = df[prefix_columns + filter_columns('1a_') + sufix_columns]
type_1b_df = df[prefix_columns + filter_columns('1b_') + sufix_columns]
type_2a_df = df[prefix_columns + filter_columns('2a_') + sufix_columns]
type_2b_df = df[prefix_columns + filter_columns('2b_') + sufix_columns]
type_3_df = df[prefix_columns + filter_columns('3_') + sufix_columns]
type_4_df = df[prefix_columns + filter_columns('4_') + sufix_columns]
type_5_df = df[prefix_columns + filter_columns('5_') + sufix_columns]
# Filter each sub DataFrame based on "ramp_type"
type_1a_df = type_1a_df[type_1a_df["ramp_type"].str.contains("Type 1A")]
type_1b_df = type_1b_df[type_1b_df["ramp_type"].str.contains("Type 1B")]
type_2a_df = type_2a_df[type_2a_df["ramp_type"].str.contains("Type 2A")]
type_2b_df = type_2b_df[type_2b_df["ramp_type"].str.contains("Type 2B")]
type_3_df = type_3_df[type_3_df["ramp_type"].str.contains("Type 3")]
type_4_df = type_4_df[type_4_df["ramp_type"].str.contains("Type 4")]
type_5_df = type_5_df[type_5_df["ramp_type"].str.contains("Type 5")]
# Reset index for each DataFrame
type_1a_df.reset_index(drop=True, inplace=True)
type_1b_df.reset_index(drop=True, inplace=True)
type_2a_df.reset_index(drop=True, inplace=True)
type_2b_df.reset_index(drop=True, inplace=True)
type_3_df.reset_index(drop=True, inplace=True)
type_4_df.reset_index(drop=True, inplace=True)
type_5_df.reset_index(drop=True, inplace=True)

# Drop the ramp_type column from each sub DataFrame
type_1a_df.drop(columns=["ramp_type"], inplace=True)
# Select the specified columns from the Type 1A DataFrame
columns_to_multiply = ['1a_c_', '1a_d_', '1a_e_', '1a_f_', '1a_cscalc', '1a_i_', '1a_j_', '1a_k_', '1a_l_', '1a_m_']
# Multiply the values of selected columns by 0.01
type_1a_df[columns_to_multiply] *= 0.01

type_1b_df.drop(columns=["ramp_type", '1b_a_r', '1b_b_r', '1b_c_r', '1b_d_r', '1b_e_r', '1b_f_r', '1b_cscalc_r', '1b_g_r', '1b_h_r', '1b_i_r', '1b_j_r', 
                         '1b_k_r', '1b_l_r', '1b_m_r', '1b_n_r', '1b_o_r', '1b_p_r', '1b_q_r', '1b_r_r', '1b_s_r'], inplace=True)
# Select the specified columns from the Type 1B DataFrame
columns_to_multiply = ['1b_c', '1b_d', '1b_e', '1b_f', '1b_cscalc', '1b_i', '1b_j', '1b_l', '1b_m', '1b_n', '1b_o', '1b_q']
# Multiply the values of selected columns by 0.01
type_1b_df[columns_to_multiply] *= 0.01

type_2a_df.drop(columns=["ramp_type"], inplace=True)
# # Select the specified columns from the Type 2A DataFrame
columns_to_multiply = ['2a_c_r', '2a_d_r', '2a_g_r', '2a_cscalc_r', '2a_j_r', '2a_k_r', '2a_m_r', '2a_n_r', '2a_o_r', '2a_p_r', '2a_s_r', '2a_t_r', '2a_u_r']
# Multiply the values of selected columns by 0.01
type_2a_df[columns_to_multiply] *= 0.01

type_2b_df.drop(columns=["ramp_type"], inplace=True)
# Select the specified columns from the Type 1B DataFrame
columns_to_multiply = ['2b_c_r', '2b_d_r', '2b_e_r', '2b_f_r', '2b_cscalc_r', '2b_h_r', '2b_k_r', '2b_l_r', '2b_o_r', '2b_r_r', '2b_s_r', '2b_v_r', '2b_w_r']
# Multiply the values of selected columns by 0.01
type_2b_df[columns_to_multiply] *= 0.01

type_3_df.drop(columns=["ramp_type", '3_corner_r'], inplace=True)
# Select the specified columns from the Type 3 DataFrame
columns_to_multiply = ['3_c_r', '3_d_r', '3_g_r', '3_j_r', '3_k_r', '3_n_r', '3_cscalc_r', '3_o_r', '3_p_r', '3_s_r', '3_t_r']
# Multiply the values of selected columns by 0.01
type_3_df[columns_to_multiply] *= 0.01

type_4_df.drop(columns=["ramp_type", "drainage", "curb_hgt", '4_st_blk_r'], inplace=True)
# Select the specified columns from the Type 4 DataFrame
columns_to_multiply = ['4_c_r', '4_d_r', '4_e_r', '4_f_r', '4_cscalc_r', '4_h_r', '4_k_r', '4_l_r', '4_o_r', '4_r_r', '4_s_r', '4_v_r']
# Multiply the values of selected columns by 0.01
type_4_df[columns_to_multiply] *= 0.01

type_5_df.drop(columns=["ramp_type"], inplace=True)
# Select the specified columns from the Type 5 DataFrame
columns_to_multiply = ['5_1_b', '5_1_c', '5_1_d', '5_1_e', '5_1_cscalc', '5_1_g', '5_1_h', '5_1_i', '5_1_j', '5_1_cscalc_2', '5_1_k2', '5_1_l2', '5_1_m', '5_1_n']
# Multiply the values of selected columns by 0.01
type_5_df[columns_to_multiply] *= 0.01

# Load your Excel template
template_path = 'Fulcrum/ADA Curb Ramp Inventory_Blank_Template.xlsx'
wb = load_workbook(template_path)

def writeToSheet(df, start_row, sheet):
    ws = wb[sheet]
    # Write DataFrame to Excel, retaining existing data and formulas
    for index, row in df.iterrows():
        for col_index, value in enumerate(row):
            cell = ws.cell(row=start_row + index, column=col_index + 1)
            # Assign value to the cell
            cell.value = value

start_row = 26
type_1a_df = type_1a_df.head(50)
writeToSheet(type_1a_df, start_row, "Type 1A")

start_row = 26
type_1b_df = type_1b_df.head(50)
writeToSheet(type_1b_df, start_row, "Type 1B")

start_row = 28
type_2a_df = type_2a_df.head(50)
writeToSheet(type_2a_df, start_row, "Type 2A")

start_row = 27
type_2b_df = type_2b_df.head(50)
writeToSheet(type_2b_df, start_row, "Type 2B")

start_row = 34
type_3_df = type_3_df.head(50)
writeToSheet(type_3_df, start_row, "Type 3")

start_row = 26
type_4_df = type_4_df.head(50)
writeToSheet(type_4_df, start_row, "Type 4")

start_row = 26
type_5_df = type_5_df.head(50)
writeToSheet(type_5_df, start_row, "Type 5")

# Save the modified workbook
wb.save(f'Fulcrum/ADA Curb Ramp Inventory_{project}.xlsx')
agsags



# Create a Pandas Excel writer using XlsxWriter as the engine
excel_file = pd.ExcelWriter("Fulcrum/ADA Curb Ramp Inventory_Blank Template2.xlsx", engine="openpyxl")

def write_df_to_xlsx_preserving_formulas(df, filepath, sheet_name="Type 1A", startrow=26):
  """
  Writes a DataFrame to an existing xlsx workbook, preserving existing data and formulas.

  Args:
      df (pandas.DataFrame): The DataFrame to write.
      filepath (str): The path to the template xlsx workbook.
      sheet_name (str, optional): The name of the sheet to write to. Defaults to "Type 1A".
      startrow (int, optional): The starting row (inclusive) for writing the DataFrame. Defaults to 26.
  """

  with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
    writer.book = pd.read_excel(filepath, sheet_name=sheet_name)  # Load existing workbook
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=0, header=False, index=False)

    # Adjust existing formulas (optional)
    worksheet = writer.sheets[sheet_name]
    for row in worksheet.iter_rows(min_row=startrow+1):  # Skip header row (if present)
      for cell in row:
        if cell.has_formula:  # Check if cell has a formula
          formula = cell.value  # Extract the formula
          cell.value = df.loc[cell.row - startrow - 1, cell.column - 1]  # Write DataFrame value
          cell.formula = formula  # Restore the formula

    writer.save()

# Example usage
filepath = "Fulcrum/ADA Curb Ramp Inventory_Blank Template2.xlsx"

write_df_to_xlsx_preserving_formulas(type_1a_df, filepath)
print("Program finished --- %s seconds ---" % (time.time() - start_time))
hghgfdhgf
# Write each DataFrame to a separate worksheet in the Excel file
type_1a_df.to_excel(excel_file, sheet_name="Type 1A", startrow=25, startcol=0, index=False)
type_1b_df.to_excel(excel_file, sheet_name="Type 1B", startrow=25, startcol=0, index=False)
type_2a_df.to_excel(excel_file, sheet_name="Type 2A", startrow=25, startcol=0, index=False)
type_2b_df.to_excel(excel_file, sheet_name="Type 2B", startrow=25, startcol=0, index=False)
type_3_df.to_excel(excel_file, sheet_name="Type 3", startrow=25, startcol=0, index=False)
type_4_df.to_excel(excel_file, sheet_name="Type 4", startrow=25, startcol=0, index=False)
type_5_df.to_excel(excel_file, sheet_name="Type 5B Median (Stop or Yield)", startrow=25, startcol=0, index=False)

# Close the Pandas Excel writer and output the Excel file
excel_file.close()
                   
# print(type_1a_df)
# print(type_1a_df.columns)
# print()
# print(type_1b_df)
# print(type_1b_df.columns)
# print()
# print(type_2a_df)
# print(type_2a_df.columns)
# print()
# print(type_2b_df)
# print(type_2b_df.columns)
# print()
# print(type_3_df)
# print(type_3_df.columns)
# print()
# print(type_4_df)
# print(type_4_df.columns)
# print()
# print(type_5_df)
# print(type_5_df.columns)

print("Program finished --- %s seconds ---" % (time.time() - start_time))