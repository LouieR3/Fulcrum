import time
import requests
import json
from fulcrum import Fulcrum
from fulcrum.exceptions import NotFoundException, InternalServerErrorException
import sys
import argparse
from datetime import date
from datetime import timedelta
import pandas as pd
from datetime import datetime, timedelta
from io import StringIO
from openpyxl import load_workbook

# https://web.fulcrumapp.com/records/636dd075-3d3a-4fd2-bab4-897700ce04a5
# https://web.fulcrumapp.com/dash/7760b3b1-8b7b-4724-a5e9-3b235f193fe7

# Get time the script started so you know how long it ran
start_time = time.time()

key = "0cff9ebe895325dcb1201a55c2b8c1b03a9e4214aaf3c4da7390e9297a300b8ccc77bce196c4abd0"

fulcrum = Fulcrum(key=key)

as_csv = fulcrum.query('SELECT * FROM "Prologis NJ PUC 2024";', 'csv')
s=str(as_csv,'utf-8')
data = StringIO(s) 
record_df=pd.read_csv(data)
# print(record_df)

# Filter and rename columns
records_filtered = record_df[['_status', '_record_id', '_title', 'number_of_warehouse_floors', 
                 'clear_height_ft', 'clear_height_in', 'building_depth_ft', 'building_depth_in', 'building_length_ft', 'building_length_in', 
                 'column_space_depth_ft', "column_space_depth_in", 'column_space_length_ft', "column_space_length_in", 'cross_dock', 
                 'speed_bay_depth_front_ft', 'speed_bay_depth_front_in', 'truck_court_depth_front_ft','truck_court_depth_front_in', 
                 'speed_bay_depth_back_ft', 'speed_bay_depth_back_in', 'truck_court_depth_back_ft', 'truck_court_depth_back_in',
                 'car_parking', 'trailer_parking', 'site_security_type', 'rail_served', 
                 'building_facade', 'floor_thickness_in_required_for_dev_only_2022_and_after', 'floor_reinforcement_required_for_dev_only_2022_and_after', 
                 'designed_floor_flatlevel_required_for_dev_only_2022_and_a', 'finished_floor_elevation_ft_required_for_dev_only_2022_a', 'general_property_notes', 
                 'main_service_transformer_kva_required_for_dev_only_2022_', 'main_service_panel_size_amps', 
                 'main_service_panel_size_volts', 'main_service_transformer_owner', 
                 "exterior_building_lighting_type", "solar_system_type", 'back_up_energy_type', 
                 'ev_car_charging', 'ev_truck_charging', 'fiber_backbone', 'green_certification']]

# Rename columns
records_filtered.columns = ['status', 'record_id', 'Property Code', 'Number of Warehouse Floors', 
                       'Clear Height Feet', 'Clear Height Inch', 'Building Depth Feet', 'Building Depth Inch', 'Building Length Feet', 'Building Length Inch', 
                       'Column Space Depth Feet', 'Column Space Depth Inch', 'Column Space Length Feet', 'Column Space Length Inch', 'Cross-Dock', 
                       'Speed Bay Depth - Front Feet', 'Speed Bay Depth - Front Inch', 'Truck Court Depth - Front Feet', 'Truck Court Depth - Front Inch', 
                       'Speed Bay Depth - Back Feet', 'Speed Bay Depth - Back Inch', 'Truck Court Depth - Back Feet', 'Truck Court Depth - Back Inch', 
                       'Car Parking', 'Trailer Parking', 'Site Security', 'Rail Served', 
                       'Building Facade', 'Floor Thickness', 'Floor Reinforcement', 
                       'Designed Floor Flat/Level', 'Finished Floor Elevation', 'Notes', 
                       'Main Service Transformer kVA', 'Main Service Size in Amps', 
                       'Main Service Size in Volts', 'Main Service Transformer Owner', 
                       'Exterior Building Lighting Type', 'Solar System', 'Back-up Energy', 
                       'EV Car Charging', 'EV Truck Charging', 'Fiber Backbone', 'Green Certification']
records_filtered.loc[:, "Property Code"] = records_filtered["Property Code"].str.split(',').str[0]
records_filtered.loc[:, "Site Security"] = records_filtered["Site Security"].fillna("")

as_csv = fulcrum.query('SELECT * FROM "Prologis NJ PUC 2024/unit_information";', 'csv')
s=str(as_csv,'utf-8')
data = StringIO(s) 
units_df=pd.read_csv(data)
# print(units_df)

# Filter and rename columns
units_df_filtered = units_df[['_child_record_id', '_record_id', '_record_status', 'unit_code', 
                              'main_floor_office_area_sq_ft', 'warehouse_area_sq_ft', 'office_mezzanine', 'office_mezzanine_nra', 'mezzanine_office_area_sq_ft', 
                              'clear_height_ft_unit', 'clear_height_in_unit', 'cross_dock_unit', 'dock_high_doors', 
                              'edge_of_dock_levelers', 'pit_levelers', 'vehicle_restraints', 
                              'drive_in_doors', 'cooling_available_in_warehouse_type',
                              'fire_suppression_system_type', 'add_calculation', 'add_k_value', 'notes_unit', 
                              'amperage_available_amps', 'kva_rights_owned', 
                              'office_lighting_type', 'warehouse_lighting_type', 
                              'smart_building_features_type']]

# Rename columns
units_df_filtered.columns = ['child_id', 'record_id', 'record_status', 'Unit Code', 
                             'Main Floor Office Area', 'Warehouse Area', 'Office Mezzanine', 'Office Mezzanine NRA', 'Mezzanine Office Area', 
                             'Clear Height Feet', 'Clear Height Inch', 'Cross-Dock', 'Dock High Doors', 
                             'Edge of Dock Levelers', 'Pit Levelers', 'Vehicle Restraints', 
                             'Drive-in Doors', 'Cooling Available in Warehouse',
                             'Fire Suppression System', 'If calculated system, add calculation', 'If ESFR, add K-value', 'Notes',
                             'Amperage available for the unit', 'KVA Rights Owned', 
                             'Office Lighting Type', 'Warehouse Lighting Type', 
                             'Smart Building features']
units_df_filtered['Unit Code'] = units_df_filtered['Unit Code'].apply(lambda x: f"Unit {x}" if isinstance(x, (int, float)) else x)
units_df_filtered.loc[:, "Smart Building features"] = units_df_filtered["Smart Building features"].fillna("")


# Filter records_filtered and units_df_filtered based on record_id
record_id = "1a173c60-0059-4ab4-9e43-7012bf1e6788"
records_filtered = records_filtered[records_filtered["record_id"] == record_id]
units_df_filtered = units_df_filtered[units_df_filtered["record_id"] == record_id]

def ordinal(n):
    """Convert an integer into its ordinal representation."""
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return str(n) + suffix

# Create and populate workbooks for each property code
for idx, record in records_filtered.iterrows():
    property_code = record['Property Code']
    record_id = record['record_id']
    print(record_id)
    
    # Load the template workbook
    wb = load_workbook('Prologis Template.xlsx')
    property_tab = wb["Property"]

    # Populate property_tab
    property_tab['E2'].value = record['Property Code']
    property_tab['E4'].value = record['Number of Warehouse Floors']
    property_tab['E5'].value = record['Clear Height Feet']
    property_tab['F5'].value = record['Clear Height Inch']
    property_tab['E6'].value = record['Building Depth Feet']
    property_tab['F6'].value = record['Building Depth Inch']
    property_tab['E7'].value = record['Building Length Feet']
    property_tab['F7'].value = record['Building Length Inch']
    property_tab['E8'].value = record['Column Space Depth Feet']
    property_tab['F8'].value = record['Column Space Depth Inch']
    property_tab['E9'].value = record['Column Space Length Feet']
    property_tab['F9'].value = record['Column Space Length Inch']
    property_tab['E10'].value = record['Cross-Dock']
    property_tab['E11'].value = record['Speed Bay Depth - Front Feet']
    property_tab['F11'].value = record['Speed Bay Depth - Front Inch']
    property_tab['E12'].value = record['Truck Court Depth - Front Feet']
    property_tab['F12'].value = record['Truck Court Depth - Front Inch']
    property_tab['E13'].value = record['Speed Bay Depth - Back Feet']
    property_tab['F13'].value = record['Speed Bay Depth - Back Inch']
    property_tab['E14'].value = record['Truck Court Depth - Back Feet']
    property_tab['F14'].value = record['Truck Court Depth - Back Inch']
    property_tab['E15'].value = record['Car Parking']
    property_tab['E16'].value = record['Trailer Parking']
    
    # Split Site Security and populate across E17-J17
    site_security_values = record['Site Security'].split(',')
    for i, value in enumerate(site_security_values):
        property_tab.cell(row=17, column=5+i).value = value.strip()
    
    property_tab['E18'].value = record['Rail Served']
    property_tab['E19'].value = record['Building Facade']
    property_tab['E21'].value = record['Floor Thickness']
    property_tab['E22'].value = record['Floor Reinforcement']
    property_tab['E23'].value = record['Designed Floor Flat/Level']
    property_tab['E24'].value = record['Finished Floor Elevation']
    property_tab['E25'].value = record['Notes']
    property_tab['E27'].value = record['Main Service Transformer kVA']
    property_tab['E28'].value = record['Main Service Size in Amps']
    property_tab['E29'].value = record['Main Service Size in Volts']
    property_tab['E31'].value = record['Main Service Transformer Owner']
    property_tab['E33'].value = record['Exterior Building Lighting Type']
    property_tab['E35'].value = record['Solar System']
    property_tab['E37'].value = record['Back-up Energy']
    property_tab['E38'].value = record['EV Car Charging']
    property_tab['E39'].value = record['EV Truck Charging']
    property_tab['E40'].value = record['Fiber Backbone']
    property_tab['E42'].value = record['Green Certification']

    print("Property Tab made")
    
    # Filter units for the current record_id and populate unit tabs
    unit_records = units_df_filtered[units_df_filtered['record_id'] == record_id]
    for unit_idx, unit_record in unit_records.iterrows():
        unit_tab_name = f"{ordinal(unit_idx+1)} Unit"
        if unit_tab_name in wb.sheetnames:
            unit_tab = wb[unit_tab_name]
            unit_tab['E2'].value = unit_record["Unit Code"]
            unit_tab['E4'].value = unit_record["Main Floor Office Area"]
            unit_tab['E5'].value = unit_record["Warehouse Area"]
            unit_tab['E6'].value = unit_record["Office Mezzanine"]
            unit_tab['E7'].value = unit_record["Mezzanine Office Area"]
            unit_tab['E9'].value = unit_record["Clear Height Feet"]
            unit_tab['F9'].value = unit_record["Clear Height Inch"]
            unit_tab['E10'].value = unit_record["Cross-Dock"]
            unit_tab['E11'].value = unit_record["Dock High Doors"]
            unit_tab['E12'].value = unit_record["Edge of Dock Levelers"]
            unit_tab['E13'].value = unit_record["Pit Levelers"]
            unit_tab['E14'].value = unit_record["Vehicle Restraints"]
            unit_tab['E15'].value = unit_record["Drive-in Doors"]
            unit_tab['E16'].value = unit_record["Cooling Available in Warehouse"]
            unit_tab['E17'].value = unit_record["Fire Suppression System"]
            unit_tab['E19'].value = unit_record["If calculated system, add calculation"]
            unit_tab['E20'].value = unit_record["If ESFR, add K-value"]
            unit_tab['E21'].value = unit_record["Notes"]
            unit_tab['E23'].value = unit_record["Amperage available for the unit"]
            unit_tab['E24'].value = unit_record["KVA Rights Owned"]
            unit_tab['E26'].value = unit_record["Office Lighting Type"]
            unit_tab['E28'].value = unit_record["Warehouse Lighting Type"]
            # Split Site Security and populate across E17-J17
            site_buildfeatures_values = unit_record['Smart Building features'].split(',')
            for i, value in enumerate(site_buildfeatures_values):
                property_tab.cell(row=30, column=5+i).value = value.strip()
            # unit_tab['E30'].value = unit_record["Smart Building features"]

    # Save the workbook with the property code as the filename
    wb.save(f"{property_code}.xlsx")

print("Program finished --- %s seconds ---" % (time.time() - start_time))